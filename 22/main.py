# python for excel using openpyxl

from openpyxl import load_workbook
# laod excel workbook

workbook = load_workbook("sample.xlsx")

# get active sheet

sheet = workbook.active

print("Sheet Title : ", sheet.title)

# read a single cell

print("A1 valeue : ",sheet["A1"].value)

# read multiple cells (row&colums)

for row in sheet.iter_rows(min_row=1,max_col=5,values_only=True):
    print(row)


# get total rows and columns

print("Total Rows : ",sheet.max_row)
print("Total Columns : ",sheet.max_column)


# read full columns (ex:column A)

for cell in sheet["A"]:
    print(cell.value)

print("End of day 22")